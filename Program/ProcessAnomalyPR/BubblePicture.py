import plotly.graph_objs as go
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
from ProcessAnomalyPR.Config import *


fig, ax = plt.subplots()


def draw_bubble(input_path):
    wb = openpyxl.load_workbook(input_path)
    sheet = wb['Sheet1']
    row_max = sheet.max_row
    col_max = sheet.max_column
    first_row_list = []
    first_col_list = []
    for col_n in range(2, col_max + 1):
        first_row_list.append(sheet.cell(row=1, column=col_n).value)
    for row_n in range(2, row_max + 1):
        first_col_list.append(sheet.cell(row=row_n, column=1).value)

    data_all = pd.read_excel(input_path)
    data_selected = data_all.loc[:, first_row_list]

    df = pd.DataFrame(data_selected)
    df.index = first_col_list
    colors = ['rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)',
              'rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)',
              'rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)','rgb(124,175,221)',]

    data = [go.Scatter(
        x=df.columns,
        y=[repo] * len(df.columns),

        mode='markers+text',
        marker=dict(
            color=colors[num],
            size=pow(df.loc[repo],0.5)*5,
            showscale=False,
        ),
        text=list(map(str, df.loc[repo])),
        textfont={
                   'size': 17,
                    'family': 'Times New Roman',
                   'color': 'rgb(0, 0, 0)',
     },
        textposition='middle center',
    )
        for num, repo in enumerate(reversed(df.index))
    ]

    layout = go.Layout(
                       plot_bgcolor='rgb(255, 255, 255)',
                       paper_bgcolor='rgb(255, 255, 255)',
                       font={
                           'size': 17,
                            'family': 'Times New Roman',
                           'color': 'rgb(0, 0, 0)',
                       },
                       width=620,
                       height=820,
                       template="ygridoff",
                       xaxis=dict(
                           # title='',
                           nticks=col_max + 1,
                           type='category',
                       ),
                       yaxis=dict(
                           range=[-1,20]
                       ),
                       showlegend=False,
                       margin=dict(l=150, r=150, t=150, b=150),
                       hovermode=False,
                       )
    fig = go.Figure(data=data, layout=layout)
    fig.write_image(f"{FIGURE_DIR}/Types_Of_Control_Flow_Anomaliesnumber_Bubble.pdf")
# py.offline.plot(fig, filename='basic-scatter.html')


if __name__ == '__main__':
    filepath = f"{SUMMARY_DIR}/repo_control_flow_anomaly_pr_num.xlsx"
    draw_bubble(filepath)
